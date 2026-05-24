from __future__ import annotations

import csv
import json
from pathlib import Path
import re
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from evolab.contracts.common import ArtifactRef
from evolab.contracts.tools import ToolResult, ToolSpec
from evolab.tools.paths import resolve_path_arguments
from evolab.tools.runtime import ToolRegistry
from evolab.tools.tables import inspect_excel_workbook, inspect_table, normalize_table, read_excel_sheet, read_table_slice


SCIENTIFIC_ARTIFACT_TOOL_NAMES = [
    "build_document_inventory",
    "discover_candidate_source_files",
    "discover_candidate_tables",
    "extract_candidate_rows",
    "build_candidate_records",
    "validate_candidate_records",
    "serialize_final_records",
]


def register_scientific_artifact_tools(
    registry: ToolRegistry,
    *,
    artifact_root: str | Path | None = None,
    base_dir: str | Path | None = None,
) -> None:
    root = Path(artifact_root) if artifact_root is not None else Path("artifacts")
    registrations = [
        (_build_document_inventory_spec(), lambda arguments: build_document_inventory(arguments, artifact_root=root)),
        (
            _discover_candidate_source_files_spec(),
            lambda arguments: discover_candidate_source_files(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("document_inventory_path", "schema_path")),
                artifact_root=root,
            ),
        ),
        (
            _discover_candidate_tables_spec(),
            lambda arguments: discover_candidate_tables(
                resolve_path_arguments(
                    arguments,
                    base_dir=base_dir,
                    names=("candidate_source_files_path", "document_inventory_path"),
                ),
                artifact_root=root,
            ),
        ),
        (
            _extract_candidate_rows_spec(),
            lambda arguments: extract_candidate_rows(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("candidate_tables_path", "schema_path")),
                artifact_root=root,
            ),
        ),
        (
            _build_candidate_records_spec(),
            lambda arguments: build_candidate_records(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("candidate_rows_path", "schema_path")),
                artifact_root=root,
            ),
        ),
        (
            _validate_candidate_records_spec(),
            lambda arguments: validate_candidate_records(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("candidate_records_path", "schema_path")),
                artifact_root=root,
            ),
        ),
        (
            _serialize_final_records_spec(),
            lambda arguments: serialize_final_records(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("records_path",)),
                artifact_root=root,
            ),
        ),
    ]
    for spec, handler in registrations:
        _register_if_missing(registry, spec, handler)


def build_document_inventory(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    root = _required_path(arguments, "root")
    if not root.exists() or not root.is_dir():
        raise ValueError(f"root is not an existing directory: {root}")
    work_item_id = _optional_string(arguments.get("work_item_id"))
    source_files = _source_files(arguments.get("source_files"), root=root)
    paths = source_files or sorted(path for path in root.rglob("*") if path.is_file())
    files = [_file_inventory(path, root=root) for path in paths if path.exists() and path.is_file()]
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "root": str(root),
        "file_count": len(files),
        "files": files,
        "summary": _inventory_summary(files),
        "warnings": [],
    }
    return _write_artifact_result(
        "build_document_inventory",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="document_inventory.json",
        artifact_kind="document_inventory",
        content=f"wrote document inventory for {len(files)} files",
    )


def discover_candidate_source_files(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    inventory = _load_json(_required_path(arguments, "document_inventory_path"))
    work_item_id = _optional_string(arguments.get("work_item_id")) or _optional_string(inventory.get("work_item_id"))
    task_terms = _task_terms(arguments.get("task_goal"), arguments.get("schema_path"))
    candidates = []
    for item in inventory.get("files", []):
        if not isinstance(item, dict):
            continue
        score, reasons = _source_score(item, task_terms)
        if score <= 0:
            continue
        candidates.append(
            {
                "path": item.get("path"),
                "relative_path": item.get("relative_path"),
                "file_type": item.get("file_type"),
                "role": item.get("role"),
                "size_bytes": item.get("size_bytes"),
                "score": round(score, 3),
                "reasons": reasons,
                "metadata": item.get("metadata", {}),
            }
        )
    candidates.sort(key=lambda item: (-float(item["score"]), str(item.get("relative_path") or item.get("path"))))
    max_sources = _optional_int(arguments.get("max_sources"), default=50)
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "document_inventory_path": str(arguments.get("document_inventory_path")),
        "candidate_sources": candidates[:max_sources],
        "source_count": min(len(candidates), max_sources),
        "scoring_policy": {
            "generic_signals": [
                "structured file type",
                "detected table shape",
                "schema/task terms in file metadata",
                "sequence-like columns or values",
            ],
            "task_terms": sorted(task_terms)[:50],
        },
    }
    return _write_artifact_result(
        "discover_candidate_source_files",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="candidate_source_files.json",
        artifact_kind="candidate_source_files",
        content=f"wrote {payload['source_count']} candidate source files",
    )


def discover_candidate_tables(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    work_item_id = _optional_string(arguments.get("work_item_id"))
    source_payload = _load_source_payload(arguments)
    work_item_id = work_item_id or _optional_string(source_payload.get("work_item_id"))
    sources = _candidate_source_items(source_payload)
    max_tables = _optional_int(arguments.get("max_tables"), default=100)
    max_sample_rows = _optional_int(arguments.get("max_sample_rows"), default=5)
    tables: list[dict[str, Any]] = []
    warnings: list[str] = []
    for source in sources:
        if len(tables) >= max_tables:
            break
        path = _source_path(source)
        if path is None or not path.exists():
            warnings.append(f"source path missing: {source.get('path')}")
            continue
        try:
            tables.extend(_tables_from_source(path, source, max_sample_rows=max_sample_rows))
        except Exception as exc:
            warnings.append(f"could not inspect {path}: {exc}")
    tables = tables[:max_tables]
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "source_ref": _source_ref(arguments),
        "candidate_tables": tables,
        "table_count": len(tables),
        "warnings": warnings,
    }
    return _write_artifact_result(
        "discover_candidate_tables",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="candidate_tables.json",
        artifact_kind="candidate_tables",
        content=f"wrote {len(tables)} candidate tables",
    )


def extract_candidate_rows(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    tables_payload = _load_json(_required_path(arguments, "candidate_tables_path"))
    work_item_id = _optional_string(arguments.get("work_item_id")) or _optional_string(tables_payload.get("work_item_id"))
    max_rows_per_table = _optional_int(arguments.get("max_rows_per_table"), default=500)
    min_sequence_length = _optional_int(arguments.get("min_sequence_length"), default=8)
    extraction_profile = _optional_string(arguments.get("sequence_extraction_profile"))
    primary_component_tables_only = bool(arguments.get("primary_component_tables_only", False))
    candidate_rows: list[dict[str, Any]] = []
    warnings: list[str] = []
    raw_tables = [table for table in tables_payload.get("candidate_tables", []) if isinstance(table, dict)]
    tables, table_selection_report = _select_tables_for_extraction(
        raw_tables,
        extraction_profile=extraction_profile,
        primary_component_tables_only=primary_component_tables_only,
    )
    for table in tables:
        if not isinstance(table, dict):
            continue
        try:
            rows = _rows_for_candidate_table(table, max_rows=max_rows_per_table)
        except Exception as exc:
            warnings.append(f"could not read rows for {table.get('source_file')}: {exc}")
            continue
        headers = _headers_for_table(table, rows)
        data_rows = _data_rows_after_header(table, rows)
        for row_offset, row in enumerate(data_rows[:max_rows_per_table]):
            normalized = _row_mapping(headers, row)
            sequence_fields = _sequence_fields(normalized, min_length=min_sequence_length)
            if not sequence_fields:
                continue
            name_fields = _name_fields(normalized, sequence_field_names={field["field"] for field in sequence_fields})
            evidence_fields = _evidence_fields(normalized, exclude={field["field"] for field in sequence_fields})
            candidate_rows.append(
                {
                    "work_item_id": work_item_id,
                    "source_file": table.get("source_file"),
                    "file_type": table.get("file_type"),
                    "sheet_name": table.get("sheet_name"),
                    "table_id": table.get("table_id"),
                    "table_index": table.get("table_index"),
                    "table_reason": table.get("reason"),
                    "table_selection": table.get("selection"),
                    "table_context": _table_context(table),
                    "row_index": _source_row_index(table, row_offset),
                    "headers": headers,
                    "row": normalized,
                    "candidate_sequence_fields": sequence_fields,
                    "candidate_label_fields": name_fields,
                    "evidence_fields": evidence_fields,
                    "confidence": _row_confidence(sequence_fields, name_fields, evidence_fields),
                    "reason": "row contains one or more plausible sequence-like values with source provenance",
                }
            )
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "candidate_tables_path": str(arguments.get("candidate_tables_path")),
        "primary_component_tables_only": primary_component_tables_only,
        "table_selection_report": table_selection_report,
        "candidate_rows": candidate_rows,
        "row_count": len(candidate_rows),
        "warnings": warnings,
    }
    return _write_artifact_result(
        "extract_candidate_rows",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="candidate_rows.json",
        artifact_kind="candidate_rows",
        content=f"wrote {len(candidate_rows)} candidate rows",
    )


def build_candidate_records(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    rows_payload = _load_json(_required_path(arguments, "candidate_rows_path"))
    work_item_id = _optional_string(arguments.get("work_item_id")) or _optional_string(rows_payload.get("work_item_id"))
    article_id = _optional_string(arguments.get("article_id")) or work_item_id
    extraction_profile = _optional_string(arguments.get("sequence_extraction_profile"))
    deduplicate_sequences = bool(arguments.get("deduplicate_sequences", True))
    records = []
    seen_sequences: set[str] = set()
    skipped_counts: dict[str, int] = {}
    for index, candidate_row in enumerate(rows_payload.get("candidate_rows", []), start=1):
        if not isinstance(candidate_row, dict):
            continue
        sequence_fields = candidate_row.get("candidate_sequence_fields")
        if not isinstance(sequence_fields, list) or not sequence_fields:
            continue
        sequence_field = _best_sequence_field(sequence_fields, extraction_profile=extraction_profile)
        if not sequence_field:
            skipped_counts["no_profile_compatible_sequence_field"] = skipped_counts.get(
                "no_profile_compatible_sequence_field", 0
            ) + 1
            continue
        sequence = _normalize_sequence(sequence_field.get("value"))
        if not sequence:
            continue
        dedup_key = _sequence_dedup_key(sequence)
        if deduplicate_sequences and dedup_key in seen_sequences:
            skipped_counts["duplicate_sequence"] = skipped_counts.get("duplicate_sequence", 0) + 1
            continue
        seen_sequences.add(dedup_key)
        component_name = _best_component_name(candidate_row, fallback=f"record-{index}")
        record = {
            "article_id": article_id,
            "work_item_id": work_item_id,
            "component_name": component_name,
            "component_type": "sequence_component",
            "sequence": sequence,
            "sequence_type": "DNA-like",
            "sequence_field": sequence_field.get("field"),
            "source_file": candidate_row.get("source_file"),
            "source_sheet": candidate_row.get("sheet_name"),
            "source_table": candidate_row.get("table_id") or candidate_row.get("table_index"),
            "source_row": candidate_row.get("row_index"),
            "evidence_text": _evidence_text(candidate_row),
            "evidence_source": _evidence_source(candidate_row),
            "source_table_context": candidate_row.get("table_context"),
            "table_selection": candidate_row.get("table_selection"),
            "acceptance_reason": _candidate_acceptance_reason(candidate_row, sequence_field, extraction_profile),
            "confidence": candidate_row.get("confidence", 0.5),
            "status": "candidate",
            "notes": _candidate_record_notes(extraction_profile),
        }
        records.append(record)
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "article_id": article_id,
        "candidate_rows_path": str(arguments.get("candidate_rows_path")),
        "sequence_extraction_profile": extraction_profile,
        "deduplicate_sequences": deduplicate_sequences,
        "skipped_counts": skipped_counts,
        "records": records,
        "record_count": len(records),
    }
    return _write_artifact_result(
        "build_candidate_records",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="candidate_records.json",
        artifact_kind="candidate_records",
        content=f"wrote {len(records)} candidate records",
    )


def _best_sequence_field(sequence_fields: list[Any], *, extraction_profile: str | None) -> dict[str, Any] | None:
    fields = [field for field in sequence_fields if isinstance(field, dict) and field.get("value")]
    if not fields:
        return None
    profile = (extraction_profile or "").casefold()
    if profile not in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return fields[0]
    scored = sorted(
        ((field, _promoter_sequence_field_score(str(field.get("field") or ""))) for field in fields),
        key=lambda item: item[1],
        reverse=True,
    )
    best, score = scored[0]
    return best if score > 0 else None


def _sequence_dedup_key(sequence: str) -> str:
    reverse = sequence.translate(str.maketrans("ACGTN", "TGCAN"))[::-1]
    return min(sequence, reverse)


def _promoter_sequence_field_score(field_name: str) -> int:
    normalized = field_name.casefold()
    score = 0
    if "promoter" in normalized and ("seq" in normalized or "sequence" in normalized):
        score += 8
    if normalized in {"promoter_sequence", "promoter_sequences", "pro_seq"}:
        score += 8
    if normalized.endswith("_sequence") or normalized.endswith("_seq"):
        score += 2
    if "pro_seq" in normalized:
        score += 5
    negative_tokens = (
        "barcode",
        "bc_",
        "pbc",
        "ibc",
        "primer",
        "adapter",
        "scaffold",
        "plasmid",
        "construct",
        "vector",
        "full",
        "oligo",
        "restriction",
        "generated",
        "predicted",
        "random",
        "motif",
    )
    if any(token in normalized for token in negative_tokens):
        score -= 12
    return score


def _candidate_acceptance_reason(
    candidate_row: dict[str, Any],
    sequence_field: dict[str, Any],
    extraction_profile: str | None,
) -> str:
    profile = (extraction_profile or "").casefold()
    field = str(sequence_field.get("field") or "")
    if profile in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return (
            f"accepted as a promoter/regulatory sequence candidate because source field {field!r} "
            "is target-compatible and row/table provenance is available"
        )
    return f"accepted as a sequence candidate because source field {field!r} has source provenance"


def _candidate_record_notes(extraction_profile: str | None) -> str:
    if (extraction_profile or "").casefold() in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return (
            "schema-constrained candidate built from a source row using reusable promoter/regulatory "
            "sequence field selection and source provenance"
        )
    return "schema-constrained candidate built from a source row with sequence-like evidence"


def validate_candidate_records(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    record_payload = _load_json(_required_path(arguments, "candidate_records_path"))
    work_item_id = _optional_string(arguments.get("work_item_id")) or _optional_string(record_payload.get("work_item_id"))
    extraction_profile = _optional_string(arguments.get("sequence_extraction_profile")) or _optional_string(
        record_payload.get("sequence_extraction_profile")
    )
    records = _records_from_payload(record_payload)
    accepted = []
    rejected = []
    for record in records:
        if not isinstance(record, dict):
            continue
        issues = _record_validation_issues(record, extraction_profile=extraction_profile)
        validated = dict(record)
        validated["validation_issues"] = issues
        if issues:
            validated["status"] = "rejected"
            rejected.append(validated)
        else:
            validated["status"] = "accepted"
            accepted.append(validated)
    payload = {
        "schema_version": "v1",
        "work_item_id": work_item_id,
        "candidate_records_path": str(arguments.get("candidate_records_path")),
        "sequence_extraction_profile": extraction_profile,
        "accepted_records": accepted,
        "rejected_records": rejected,
        "record_count": len(records),
        "accepted_count": len(accepted),
        "rejected_count": len(rejected),
        "validation_policy": [
            "record has non-empty source provenance",
            "record fields are not empty",
            "sequence fields are plausible DNA-like strings when present",
            "source file exists",
            "target-component semantics are present for profile-constrained extraction",
            "prediction-only, primer, scaffold, plasmid, construct, barcode, and assay-detail artifacts are rejected or staged",
        ],
    }
    return _write_artifact_result(
        "validate_candidate_records",
        payload,
        artifact_root=artifact_root,
        work_item_id=work_item_id,
        filename="validated_records.json",
        artifact_kind="validated_records",
        content=f"validated {len(records)} records ({len(accepted)} accepted)",
    )


def serialize_final_records(arguments: dict[str, Any], *, artifact_root: Path | None = None) -> ToolResult:
    records_path_value = arguments.get("records_path")
    if isinstance(records_path_value, str) and records_path_value:
        payload = _load_json(Path(records_path_value).expanduser())
        records = _accepted_records_from_payload(payload)
    else:
        records = arguments.get("records", [])
    if not isinstance(records, list) or not all(isinstance(record, dict) for record in records):
        raise ValueError("records must be a list of objects or records_path must point to a supported record artifact")
    artifact_name = str(arguments.get("artifact_name") or "biology_component_records.jsonl")
    work_item_id = _optional_string(arguments.get("work_item_id"))
    root = artifact_root or Path("artifacts")
    output_path = _artifact_path(root, filename=artifact_name, work_item_id=work_item_id)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    _write_jsonl(output_path, records)
    artifact_refs = [
        ArtifactRef(
            uri=str(output_path),
            type="dataset",
            metadata={
                "filename": output_path.name,
                "artifact_kind": "final_records",
                "record_count": len(records),
                "format": "jsonl",
                **({"work_item_id": work_item_id} if work_item_id else {}),
            },
        )
    ]
    if bool(arguments.get("also_write_final_records", False)) and output_path.name != "final_records.jsonl":
        final_path = _artifact_path(root, filename="final_records.jsonl", work_item_id=work_item_id)
        final_path.parent.mkdir(parents=True, exist_ok=True)
        _write_jsonl(final_path, records)
        artifact_refs.append(
            ArtifactRef(
                uri=str(final_path),
                type="dataset",
                metadata={
                    "filename": final_path.name,
                    "artifact_kind": "final_records",
                    "record_count": len(records),
                    "format": "jsonl",
                    **({"work_item_id": work_item_id} if work_item_id else {}),
                },
            )
        )
    return ToolResult(
        call_id="serialize_final_records",
        status="ok",
        content=f"serialized {len(records)} final records",
        artifact_refs=artifact_refs,
        metadata={"record_count": len(records), "paths": [ref.uri for ref in artifact_refs]},
    )


def _file_inventory(path: Path, *, root: Path) -> dict[str, Any]:
    suffix = path.suffix.casefold()
    metadata: dict[str, Any] = {}
    file_type = _file_type(path)
    if file_type == "spreadsheet":
        try:
            workbook = _inspect_workbook_lightweight(path)
            metadata.update(
                {
                    "sheet_count": workbook.get("sheet_count", 0),
                    "sheets": workbook.get("sheets", []),
                }
            )
        except Exception as exc:
            metadata["inspection_error"] = str(exc)
    elif file_type in {"csv_tsv", "markdown"}:
        try:
            inspected = inspect_table({"path": str(path), "max_rows_preview": 5}).metadata
            metadata.update(
                {
                    "row_count": inspected.get("row_count", 0),
                    "column_count": inspected.get("column_count", 0),
                    "headers": inspected.get("headers", []),
                    "preview_rows": inspected.get("preview_rows", []),
                }
            )
            if inspected.get("plain_text_table_block"):
                metadata["plain_text_table_block"] = inspected.get("plain_text_table_block")
        except Exception as exc:
            metadata["inspection_error"] = str(exc)
    elif suffix in {".md", ".txt", ".json"}:
        try:
            text = path.read_text(encoding="utf-8")
            metadata.update({"char_count": len(text), "sample": text[:500]})
        except UnicodeDecodeError:
            metadata["inspection_error"] = "not utf-8 text"
    stat = path.stat()
    return {
        "path": str(path),
        "relative_path": _relative_path(path, root),
        "filename": path.name,
        "suffix": path.suffix,
        "file_type": file_type,
        "role": _file_role(path),
        "size_bytes": stat.st_size,
        "metadata": metadata,
        "excluded_as_prior_output": path.name in _prior_output_filenames(),
    }


def _inventory_summary(files: list[dict[str, Any]]) -> dict[str, int]:
    summary: dict[str, int] = {}
    for item in files:
        key = str(item.get("file_type") or "unknown")
        summary[key] = summary.get(key, 0) + 1
    return summary


def _file_type(path: Path) -> str:
    suffix = path.suffix.casefold()
    if path.name == "main_text.md":
        return "markdown"
    if suffix in {".xlsx", ".xlsm", ".xls"}:
        return "spreadsheet"
    if suffix in {".csv", ".tsv"}:
        return "csv_tsv"
    if suffix in {".md", ".markdown"}:
        return "markdown"
    if suffix in {".txt", ".text"}:
        return "text"
    if suffix == ".json":
        return "json"
    if suffix in {".pdf"}:
        return "pdf"
    return "other"


def _file_role(path: Path) -> str:
    if path.name == "main_text.md":
        return "main_text"
    if path.name == "manifest.json":
        return "manifest"
    parts = {part.casefold() for part in path.parts}
    name = path.name.casefold()
    if "supplementary" in parts or name.startswith(("sup", "supp", "sub", "table_s", "dataset_s")):
        return "supplementary"
    if path.name in _prior_output_filenames():
        return "prior_output"
    return "source"


def _prior_output_filenames() -> set[str]:
    return {"biology_component_records.jsonl", "biology_component_report.md", "biology_component_report.json"}


def _source_score(item: dict[str, Any], task_terms: set[str]) -> tuple[float, list[str]]:
    if item.get("excluded_as_prior_output"):
        return 0.0, ["excluded prior output artifact"]
    file_type = str(item.get("file_type") or "")
    role = str(item.get("role") or "")
    metadata = item.get("metadata") if isinstance(item.get("metadata"), dict) else {}
    score = 0.0
    reasons: list[str] = []
    if file_type == "spreadsheet":
        score += 6
        reasons.append("spreadsheet source")
    elif file_type == "csv_tsv":
        score += 5
        reasons.append("delimited table source")
    elif file_type == "markdown":
        score += 3
        reasons.append("markdown/text source")
    elif file_type == "text":
        score += 1
        reasons.append("text source")
    if role == "supplementary":
        score += 2
        reasons.append("supplementary source")
    if role == "main_text":
        score += 1
        reasons.append("main text source")
    table_shape = _metadata_table_shape(metadata)
    if table_shape:
        score += min(4, table_shape)
        reasons.append("detected table shape")
    text = json.dumps(metadata, sort_keys=True).casefold()
    matched_terms = sorted(term for term in task_terms if term and term in text)[:10]
    if matched_terms:
        score += min(3, 0.5 * len(matched_terms))
        reasons.append(f"matched task/schema terms: {', '.join(matched_terms)}")
    if _metadata_has_sequence_signal(metadata):
        score += 3
        reasons.append("sequence-like headers or values")
    return score, reasons


def _metadata_table_shape(metadata: dict[str, Any]) -> float:
    sheet_count = metadata.get("sheet_count")
    if isinstance(sheet_count, int) and sheet_count > 0:
        return float(sheet_count)
    row_count = metadata.get("row_count")
    column_count = metadata.get("column_count")
    if isinstance(row_count, int) and isinstance(column_count, int) and row_count > 0 and column_count > 0:
        return 1.0
    return 0.0


def _metadata_has_sequence_signal(metadata: dict[str, Any]) -> bool:
    text = json.dumps(metadata, sort_keys=True)
    if any(token in text.casefold() for token in ("sequence", "seq", "dna", "nucleotide")):
        return True
    return any(_is_dna_like(str(value), min_length=12) for value in _iter_leaf_values(metadata))


def _tables_from_source(path: Path, source: dict[str, Any], *, max_sample_rows: int) -> list[dict[str, Any]]:
    file_type = str(source.get("file_type") or _file_type(path))
    if file_type == "spreadsheet":
        source_metadata = source.get("metadata") if isinstance(source.get("metadata"), dict) else {}
        source_sheets = source_metadata.get("sheets") if isinstance(source_metadata, dict) else None
        workbook = (
            {"path": str(path), "sheet_count": len(source_sheets), "sheets": source_sheets}
            if isinstance(source_sheets, list) and source_sheets
            else _inspect_workbook_lightweight(path)
        )
        skip_samples = path.stat().st_size > 5_000_000
        tables = []
        for index, sheet in enumerate(workbook.get("sheets", [])):
            if not isinstance(sheet, dict):
                continue
            sheet_name = str(sheet.get("sheet_name") or "")
            sample_rows = (
                []
                if skip_samples
                else _read_excel_rows_limited(path, sheet_name=sheet_name, max_rows=max_sample_rows + 1)
            )
            tables.append(
                {
                    "table_id": f"{path.name}::{sheet_name}",
                    "source_file": str(path),
                    "file_type": file_type,
                    "sheet_name": sheet_name,
                    "table_index": index,
                    "header_row_index": sheet.get("header_row_index") if isinstance(sheet.get("header_row_index"), int) else 0,
                    "headers": _string_list(sheet.get("headers")),
                    "row_count": sheet.get("non_empty_rows") or sheet.get("max_row"),
                    "column_count": sheet.get("max_column"),
                    "sample_rows": sample_rows[: max_sample_rows + 1],
                    "sample_rows_omitted": skip_samples,
                    "reason": _table_reason(sheet.get("headers")),
                }
            )
        return tables
    if file_type == "csv_tsv":
        inspected = inspect_table({"path": str(path), "max_rows_preview": max_sample_rows + 1}).metadata
        return [
            {
                "table_id": f"{path.name}::table-0",
                "source_file": str(path),
                "file_type": file_type,
                "sheet_name": None,
                "table_index": 0,
                "header_row_index": 0,
                "headers": inspected.get("headers", []),
                "row_count": inspected.get("row_count"),
                "column_count": inspected.get("column_count"),
                "sample_rows": inspected.get("preview_rows", []),
                "reason": _table_reason(inspected.get("headers")),
            }
        ]
    if file_type == "markdown":
        markdown_tables = _markdown_tables(path, max_sample_rows=max_sample_rows)
        if markdown_tables:
            return markdown_tables
        inspected = inspect_table({"path": str(path), "max_rows_preview": max_sample_rows + 1}).metadata
        rows = inspected.get("preview_rows", [])
        if not rows:
            return []
        return [
            {
                "table_id": f"{path.name}::table-0",
                "source_file": str(path),
                "file_type": file_type,
                "sheet_name": None,
                "table_index": 0,
                "header_row_index": 0,
                "headers": inspected.get("headers", []),
                "row_count": inspected.get("row_count"),
                "column_count": inspected.get("column_count"),
                "sample_rows": rows,
                "rows": rows,
                "plain_text_table_block": inspected.get("plain_text_table_block"),
                "reason": _table_reason(inspected.get("headers")),
            }
        ]
    return []


def _markdown_tables(path: Path, *, max_sample_rows: int) -> list[dict[str, Any]]:
    lines = path.read_text(encoding="utf-8").splitlines()
    sheet_index_labels = _markdown_sheet_index_labels(lines)
    tables: list[dict[str, Any]] = []
    index = 0
    current_heading = ""
    while index < len(lines):
        stripped = lines[index].strip()
        if stripped.startswith("#"):
            current_heading = stripped.lstrip("#").strip()
        if not (stripped.startswith("|") and stripped.endswith("|")):
            index += 1
            continue
        start = index
        block = []
        while index < len(lines):
            current = lines[index].strip()
            if not (current.startswith("|") and current.endswith("|")):
                break
            cells = [cell.strip() for cell in current.strip("|").split("|")]
            if cells and all(set(cell) <= {"-", ":", " "} for cell in cells):
                index += 1
                continue
            block.append(cells)
            index += 1
        if block:
            headers = [_normalize_header(cell) for cell in block[0]]
            context_start = max(0, start - 3)
            context_end = min(len(lines), index + 3)
            section_label = _sheet_label_for_heading(current_heading, sheet_index_labels)
            tables.append(
                {
                    "table_id": f"{path.name}::table-{len(tables)}",
                    "source_file": str(path),
                    "file_type": "markdown",
                    "sheet_name": None,
                    "table_index": len(tables),
                    "header_row_index": 0,
                    "headers": headers,
                    "row_count": max(0, len(block) - 1),
                    "column_count": len(headers),
                    "sample_rows": block[: max_sample_rows + 1],
                    "rows": block,
                    "surrounding_context": "\n".join(lines[context_start:context_end])[:1000],
                    "section_heading": current_heading,
                    "section_label": section_label,
                    "reason": _table_reason(headers),
                }
            )
    return tables


def _markdown_sheet_index_labels(lines: list[str]) -> dict[str, str]:
    labels: dict[str, str] = {}
    index = 0
    while index < len(lines):
        stripped = lines[index].strip()
        if not (stripped.startswith("|") and stripped.endswith("|")):
            index += 1
            continue
        block = []
        while index < len(lines):
            current = lines[index].strip()
            if not (current.startswith("|") and current.endswith("|")):
                break
            cells = [cell.strip() for cell in current.strip("|").split("|")]
            if cells and all(set(cell) <= {"-", ":", " "} for cell in cells):
                index += 1
                continue
            block.append(cells)
            index += 1
        if len(block) < 2:
            continue
        headers = [_normalize_header(cell) for cell in block[0]]
        if "dataset" not in headers or "sheet" not in headers:
            continue
        dataset_index = headers.index("dataset")
        sheet_index = headers.index("sheet")
        for row in block[1:]:
            if len(row) <= max(dataset_index, sheet_index):
                continue
            sheet = str(row[sheet_index]).strip()
            dataset = str(row[dataset_index]).strip()
            if sheet and dataset:
                labels[sheet] = dataset
    return labels


def _sheet_label_for_heading(heading: str, labels: dict[str, str]) -> str | None:
    match = re.search(r"\bsheet\s+([A-Za-z0-9_.-]+)", heading, flags=re.I)
    if not match:
        return None
    return labels.get(match.group(1))


def _inspect_workbook_lightweight(path: Path) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return inspect_excel_workbook({"path": str(path)}).metadata
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return inspect_excel_workbook({"path": str(path)}).metadata
    sheets = []
    try:
        for sheet in workbook.worksheets:
            sample_rows = _sheet_rows_limited(sheet, max_rows=20)
            header_result = _detect_header_lightweight(sample_rows)
            sheets.append(
                {
                    "sheet_name": sheet.title,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                    "non_empty_rows": sheet.max_row,
                    "header_row_index": header_result["header_row_index"],
                    "headers": header_result["headers"],
                }
            )
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()
    return {"path": str(path), "sheet_count": len(sheets), "sheets": sheets}


def _read_excel_rows_limited(path: Path, *, sheet_name: Any, max_rows: int) -> list[list[Any]]:
    if path.suffix.casefold() in {".xlsx", ".xlsm"}:
        rows = _read_xlsx_rows_limited_from_zip(path, sheet_name=sheet_name, max_rows=max_rows)
        if rows is not None:
            return rows
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        result = read_excel_sheet({"path": str(path), "sheet_name": sheet_name, "start_row": 0, "end_row": max_rows})
        return [_list_row(row) for row in result.metadata.get("rows", [])]
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        if isinstance(sheet_name, str) and sheet_name:
            sheet = workbook[sheet_name]
        else:
            sheet = workbook.worksheets[0]
        return _sheet_rows_limited(sheet, max_rows=max_rows)
    finally:
        close = getattr(workbook, "close", None)
        if callable(close):
            close()


def _read_xlsx_rows_limited_from_zip(path: Path, *, sheet_name: Any, max_rows: int) -> list[list[Any]] | None:
    try:
        with ZipFile(path) as archive:
            shared_strings = _xlsx_shared_strings(archive)
            targets = _xlsx_sheet_targets(archive)
            if not targets:
                return None
            target = _select_xlsx_target(targets, sheet_name)
            if target is None or target not in archive.namelist():
                return None
            return _xlsx_sheet_rows_limited(archive, target, shared_strings, max_rows=max_rows)
    except Exception:
        return None


def _xlsx_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return ["".join(text.text or "" for text in item.findall(".//{*}t")) for item in root.findall(".//{*}si")]


def _xlsx_sheet_targets(archive: ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_by_id = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("{*}Relationship")
        if "Id" in rel.attrib and "Target" in rel.attrib
    }
    targets = []
    for sheet in workbook.findall(".//{*}sheet"):
        name = sheet.attrib.get("name", "Sheet")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rel_by_id.get(str(rel_id))
        if target:
            targets.append((name, _xlsx_target_path(target)))
    return targets


def _select_xlsx_target(targets: list[tuple[str, str]], sheet_name: Any) -> str | None:
    if isinstance(sheet_name, str) and sheet_name:
        for name, target in targets:
            if name == sheet_name:
                return target
    return targets[0][1] if targets else None


def _xlsx_target_path(target: str) -> str:
    normalized = target.lstrip("/")
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def _xlsx_sheet_rows_limited(
    archive: ZipFile,
    target: str,
    shared_strings: list[str],
    *,
    max_rows: int,
) -> list[list[Any]]:
    rows: list[list[Any]] = []
    with archive.open(target) as handle:
        for _, elem in ET.iterparse(handle, events=("end",)):
            if not elem.tag.endswith("row"):
                continue
            row_values: list[Any] = []
            for cell in elem:
                if not cell.tag.endswith("c"):
                    continue
                index = _xlsx_cell_column_index(cell.attrib.get("r", ""))
                while len(row_values) <= index:
                    row_values.append("")
                row_values[index] = _xlsx_cell_value(cell, shared_strings)
            rows.append(row_values)
            elem.clear()
            if len(rows) >= max_rows:
                break
    return rows


def _xlsx_cell_column_index(reference: str) -> int:
    letters = "".join(char for char in reference if char.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(0, index - 1)


def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    if cell.attrib.get("t") == "inlineStr":
        return "".join(text.text or "" for text in cell.findall(".//{*}t"))
    value = cell.find("{*}v")
    if value is None:
        return ""
    text = value.text or ""
    if cell.attrib.get("t") == "s":
        try:
            return shared_strings[int(text)]
        except (ValueError, IndexError):
            return text
    return text


def _select_tables_for_extraction(
    tables: list[dict[str, Any]],
    *,
    extraction_profile: str | None,
    primary_component_tables_only: bool,
) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    classified = []
    for table in tables:
        selection = _table_selection_signals(table, extraction_profile=extraction_profile)
        enriched = dict(table)
        enriched["selection"] = selection
        classified.append(enriched)
    if not primary_component_tables_only:
        return classified, {
            "enabled": False,
            "input_table_count": len(tables),
            "selected_table_count": len(classified),
        }

    selected: list[dict[str, Any]] = []
    rejected: list[dict[str, Any]] = []
    by_source: dict[str, list[dict[str, Any]]] = {}
    for table in classified:
        source = str(table.get("source_file") or "")
        by_source.setdefault(source, []).append(table)
    for source_tables in by_source.values():
        candidates = [table for table in source_tables if _table_is_extractable_primary_candidate(table)]
        if not candidates:
            rejected.extend(source_tables)
            continue
        candidates.sort(
            key=lambda table: (
                -float(table.get("selection", {}).get("score", 0)),
                int(table.get("table_index") if isinstance(table.get("table_index"), int) else 10_000),
                str(table.get("table_id") or ""),
            )
        )
        selected.append(candidates[0])
        selected_ids = {id(candidates[0])}
        rejected.extend(table for table in source_tables if id(table) not in selected_ids)
    return selected, {
        "enabled": True,
        "input_table_count": len(tables),
        "selected_table_count": len(selected),
        "selected_tables": [table.get("table_id") for table in selected],
        "rejected_tables": [
            {
                "table_id": table.get("table_id"),
                "reasons": table.get("selection", {}).get("reasons", []),
                "score": table.get("selection", {}).get("score"),
            }
            for table in rejected
        ],
        "policy": (
            "precision-oriented primary component table selection; keeps the strongest primary "
            "component table per source and stages secondary assay/design/detail tables"
        ),
    }


def _table_is_extractable_primary_candidate(table: dict[str, Any]) -> bool:
    selection = table.get("selection") if isinstance(table.get("selection"), dict) else {}
    if selection.get("hard_reject"):
        return False
    return float(selection.get("score", 0) or 0) > 0


def _table_selection_signals(table: dict[str, Any], *, extraction_profile: str | None) -> dict[str, Any]:
    headers = _string_list(table.get("headers"))
    header_text = " ".join(headers).casefold()
    context = _table_context(table)
    context_text = json.dumps(context, sort_keys=True, default=str).casefold()
    profile = (extraction_profile or "").casefold()
    reasons: list[str] = []
    score = 0.0
    hard_reject = False

    target_sequence_headers = _target_sequence_headers(headers, extraction_profile=profile)
    sequence_headers = [header for header in headers if any(token in header.casefold() for token in ("sequence", "seq", "dna"))]
    if target_sequence_headers:
        score += 10
        reasons.append(f"target-compatible sequence headers: {', '.join(target_sequence_headers[:5])}")
    elif sequence_headers:
        score += 3
        reasons.append(f"generic sequence headers: {', '.join(sequence_headers[:5])}")

    if any(token in header_text for token in ("name", "id", "component", "part", "variant")):
        score += 2
        reasons.append("label/name/id columns")
    if any(token in context_text or token in header_text for token in ("promoter", "regulatory", "sigma", "u6", "polii", "pol_ii")):
        score += 2
        reasons.append("target component terms in table headers or context")
    if any(
        token in header_text
        for token in ("activity", "score", "measurement", "transcription", "rna_count", "dna_count", "fluorescence", "od600")
    ):
        score += 1
        reasons.append("experimental measurement columns")

    if _table_is_prediction_only(table):
        hard_reject = True
        reasons.append("prediction-only/generated-only table")
    if _table_has_non_target_sequence_only(headers, extraction_profile=profile):
        hard_reject = True
        reasons.append("sequence columns describe non-target artifacts")
    if _table_is_barcode_replicate_detail(table):
        score -= 12
        reasons.append("barcode/internal-barcode replicate assay detail table")
    if any(token in context_text for token in ("primer", "adapter", "linker", "restriction site", "plasmid", "construct")):
        score -= 8
        reasons.append("negative artifact context")
    return {
        "score": round(score, 3),
        "hard_reject": hard_reject,
        "reasons": reasons,
        "context": context,
    }


def _target_sequence_headers(headers: list[str], *, extraction_profile: str) -> list[str]:
    if extraction_profile in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return [
            header
            for header in headers
            if _promoter_sequence_field_score(header) > 0 and not _header_is_non_target_sequence_artifact(header)
        ]
    return [header for header in headers if any(token in header.casefold() for token in ("sequence", "seq", "dna"))]


def _table_is_prediction_only(table: dict[str, Any]) -> bool:
    text = _table_text(table)
    prediction_terms = ("generated_sequences", "generated sequence", "predicted_value", "predicted-only", "prediction-only")
    measurement_terms = ("measured", "activity", "transcription", "rna_count", "dna_count", "edit_score", "od600")
    return any(term in text for term in prediction_terms) and not any(term in text for term in measurement_terms)


def _table_has_non_target_sequence_only(headers: list[str], *, extraction_profile: str) -> bool:
    sequence_headers = [header for header in headers if any(token in header.casefold() for token in ("sequence", "seq", "dna"))]
    if not sequence_headers:
        return False
    if extraction_profile in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return not _target_sequence_headers(headers, extraction_profile=extraction_profile)
    return all(_header_is_non_target_sequence_artifact(header) for header in sequence_headers)


def _header_is_non_target_sequence_artifact(header: str) -> bool:
    normalized = header.casefold()
    negative_tokens = (
        "primer",
        "barcode",
        "adapter",
        "linker",
        "scaffold",
        "plasmid",
        "construct",
        "vector",
        "oligo",
        "pbc_seq",
        "ibc_seq",
        "restriction",
    )
    return any(token in normalized for token in negative_tokens) and "promoter" not in normalized


def _table_is_barcode_replicate_detail(table: dict[str, Any]) -> bool:
    text = _table_text(table)
    return (
        ("pbc_seq" in text or "ibc_seq" in text)
        and ("replicate" in text or "read_count" in text or "freq" in text or "raw_" in text)
    )


def _table_text(table: dict[str, Any]) -> str:
    values = [
        table.get("headers"),
        table.get("sheet_name"),
        table.get("table_id"),
        table.get("reason"),
        table.get("surrounding_context"),
        table.get("section_heading"),
        table.get("section_label"),
    ]
    return " ".join(json.dumps(value, sort_keys=True, default=str) for value in values if value).casefold()


def _table_context(table: dict[str, Any]) -> dict[str, Any]:
    context = {
        "table_id": table.get("table_id"),
        "sheet_name": table.get("sheet_name"),
        "table_index": table.get("table_index"),
        "reason": table.get("reason"),
        "section_heading": table.get("section_heading"),
        "section_label": table.get("section_label"),
        "surrounding_context": table.get("surrounding_context"),
    }
    return {key: value for key, value in context.items() if value not in (None, "", [])}


def _sheet_rows_limited(sheet: Any, *, max_rows: int) -> list[list[Any]]:
    rows: list[list[Any]] = []
    for row in sheet.iter_rows(max_row=max_rows, values_only=True):
        rows.append([cell for cell in row])
    return rows


def _detect_header_lightweight(rows: list[list[Any]]) -> dict[str, Any]:
    best_index: int | None = None
    best_score = -1.0
    for index, row in enumerate(rows[:20]):
        non_empty = [cell for cell in row if cell is not None and str(cell).strip()]
        string_count = sum(isinstance(cell, str) for cell in non_empty)
        score = string_count + min(len(non_empty), 4) * 0.25
        if len(non_empty) <= 1:
            score -= 2.0
        if score > best_score:
            best_index = index
            best_score = score
    if best_index is None:
        return {"header_row_index": None, "headers": []}
    return {"header_row_index": best_index, "headers": [_normalize_header(cell) for cell in rows[best_index]]}


def _rows_for_candidate_table(table: dict[str, Any], *, max_rows: int) -> list[list[Any]]:
    file_type = str(table.get("file_type") or "")
    source_file = _source_file_from_table(table)
    if source_file is None:
        return []
    if file_type == "spreadsheet":
        return _read_excel_rows_limited(source_file, sheet_name=table.get("sheet_name"), max_rows=max_rows + 1)
    if file_type == "csv_tsv":
        result = read_table_slice({"path": str(source_file), "start_row": 0, "end_row": max_rows + 1})
        return [_list_row(row) for row in result.metadata.get("rows", [])]
    rows = table.get("rows") or table.get("sample_rows") or []
    return [_list_row(row) for row in rows[: max_rows + 1]]


def _headers_for_table(table: dict[str, Any], rows: list[list[Any]]) -> list[str]:
    headers = _string_list(table.get("headers"))
    if headers:
        return headers
    normalized = normalize_table({"rows": rows}).metadata
    return _string_list(normalized.get("headers"))


def _data_rows_after_header(table: dict[str, Any], rows: list[list[Any]]) -> list[list[Any]]:
    header_row_index = table.get("header_row_index")
    if isinstance(header_row_index, int):
        return rows[header_row_index + 1 :]
    return rows[1:] if rows else []


def _row_mapping(headers: list[str], row: list[Any]) -> dict[str, str]:
    if not headers:
        headers = [f"column_{index + 1}" for index in range(len(row))]
    return {
        headers[index] if index < len(headers) else f"column_{index + 1}": "" if value is None else str(value).strip()
        for index, value in enumerate(row)
    }


def _sequence_fields(row: dict[str, str], *, min_length: int) -> list[dict[str, str]]:
    fields = []
    for field, value in row.items():
        sequence = _normalize_sequence(value)
        if sequence and _is_dna_like(sequence, min_length=min_length):
            fields.append({"field": field, "value": sequence})
    return fields


def _name_fields(row: dict[str, str], *, sequence_field_names: set[str]) -> list[dict[str, str]]:
    fields = []
    preferred = ("name", "label", "id", "component", "construct", "part", "variant")
    for field, value in row.items():
        if field in sequence_field_names or not value:
            continue
        field_text = field.casefold()
        if any(token in field_text for token in preferred):
            fields.append({"field": field, "value": value})
    if not fields:
        for field, value in row.items():
            if field not in sequence_field_names and value and not _is_dna_like(value, min_length=8):
                fields.append({"field": field, "value": value})
                break
    return fields


def _evidence_fields(row: dict[str, str], *, exclude: set[str]) -> list[dict[str, str]]:
    fields = []
    for field, value in row.items():
        if field in exclude or not value:
            continue
        if any(token in field.casefold() for token in ("evidence", "method", "activity", "score", "context", "note")):
            fields.append({"field": field, "value": value})
    return fields


def _row_confidence(
    sequence_fields: list[dict[str, str]],
    name_fields: list[dict[str, str]],
    evidence_fields: list[dict[str, str]],
) -> float:
    score = 0.45
    if sequence_fields:
        score += 0.25
    if name_fields:
        score += 0.15
    if evidence_fields:
        score += 0.15
    return round(min(1.0, score), 3)


def _record_validation_issues(record: dict[str, Any], *, extraction_profile: str | None = None) -> list[str]:
    issues = []
    if not record.get("source_file"):
        issues.append("missing source_file")
    elif not Path(str(record["source_file"])).exists():
        issues.append("source_file does not exist")
    if not record.get("component_name"):
        issues.append("missing component_name")
    sequence = _normalize_sequence(record.get("sequence"))
    if record.get("sequence") and not _is_dna_like(sequence, min_length=8):
        issues.append("sequence is not DNA-like")
    if not record.get("evidence_source"):
        issues.append("missing evidence_source")
    if not record.get("acceptance_reason"):
        issues.append("missing acceptance_reason")
    issues.extend(_profile_validation_issues(record, extraction_profile=extraction_profile))
    return issues


def _profile_validation_issues(record: dict[str, Any], *, extraction_profile: str | None) -> list[str]:
    profile = (extraction_profile or "").casefold()
    if profile not in {"promoter", "promoter_sequence", "regulatory_sequence"}:
        return []
    text = _record_semantic_text(record)
    sequence_field = str(record.get("sequence_field") or "").casefold()
    target_terms = ("promoter", "regulatory", "sigma", "u6", "polii", "pol_ii")
    if not any(term in text for term in target_terms):
        return ["missing target-component semantics for promoter profile"]
    negative_sequence_fields = (
        "primer",
        "barcode",
        "adapter",
        "linker",
        "scaffold",
        "plasmid",
        "construct",
        "vector",
        "oligo",
        "pbc_seq",
        "ibc_seq",
        "full_oligo",
        "restriction",
    )
    if any(token in sequence_field for token in negative_sequence_fields) and "promoter" not in sequence_field:
        return [f"sequence field is a non-target artifact field: {sequence_field}"]
    if _is_prediction_only_record(record):
        return ["prediction-only or generated-only sequence without experimental component evidence"]
    if _is_barcode_replicate_detail_record(record):
        return ["barcode/replicate assay-detail row is not a primary component record"]
    return []


def _record_semantic_text(record: dict[str, Any]) -> str:
    values = [
        record.get("component_name"),
        record.get("component_type"),
        record.get("sequence_field"),
        record.get("evidence_text"),
        record.get("evidence_source"),
        record.get("source_table_context"),
        record.get("table_selection"),
        record.get("acceptance_reason"),
        record.get("notes"),
    ]
    return " ".join(json.dumps(value, sort_keys=True, default=str) for value in values if value).casefold()


def _is_prediction_only_record(record: dict[str, Any]) -> bool:
    text = _record_semantic_text(record)
    prediction_terms = ("generated_sequences", "generated sequence", "predicted_value", "prediction-only", "predicted-only")
    measurement_terms = (
        "measured",
        "activity",
        "transcription",
        "rna_count",
        "dna_count",
        "edit_score",
        "od600",
        "fluorescence",
    )
    return any(term in text for term in prediction_terms) and not any(term in text for term in measurement_terms)


def _is_barcode_replicate_detail_record(record: dict[str, Any]) -> bool:
    text = _record_semantic_text(record)
    return (
        ("pbc_seq" in text or "ibc_seq" in text)
        and ("replicate" in text or "read_count" in text or "freq" in text or "raw_" in text)
    )


def _best_component_name(candidate_row: dict[str, Any], *, fallback: str) -> str:
    fields = candidate_row.get("candidate_label_fields")
    if isinstance(fields, list):
        for field in fields:
            if isinstance(field, dict) and field.get("value"):
                return str(field["value"])
    row = candidate_row.get("row")
    if isinstance(row, dict):
        for value in row.values():
            if isinstance(value, str) and value and not _is_dna_like(value, min_length=8):
                return value
    return fallback


def _evidence_text(candidate_row: dict[str, Any]) -> str:
    row = candidate_row.get("row")
    if isinstance(row, dict):
        return json.dumps(row, sort_keys=True)
    return ""


def _evidence_source(candidate_row: dict[str, Any]) -> dict[str, Any]:
    source: dict[str, Any] = {
        "path": candidate_row.get("source_file"),
        "sheet_name": candidate_row.get("sheet_name"),
        "table_id": candidate_row.get("table_id") or candidate_row.get("table_index"),
        "row_index": candidate_row.get("row_index"),
    }
    table_context = candidate_row.get("table_context")
    if isinstance(table_context, dict):
        source["table_context"] = table_context
    return {key: value for key, value in source.items() if value not in (None, "")}


def _accepted_records_from_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    accepted = payload.get("accepted_records")
    if isinstance(accepted, list):
        return [item for item in accepted if isinstance(item, dict)]
    records = payload.get("records")
    if isinstance(records, list):
        return [item for item in records if isinstance(item, dict) and item.get("status") != "rejected"]
    return []


def _records_from_payload(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, list):
        return [item for item in payload if isinstance(item, dict)]
    if not isinstance(payload, dict):
        return []
    for key in ("records", "accepted_records", "candidate_records"):
        value = payload.get(key)
        if isinstance(value, list):
            return [item for item in value if isinstance(item, dict)]
    return []


def _write_artifact_result(
    call_id: str,
    payload: dict[str, Any],
    *,
    artifact_root: Path | None,
    work_item_id: str | None,
    filename: str,
    artifact_kind: str,
    content: str,
) -> ToolResult:
    root = artifact_root or Path("artifacts")
    path = _artifact_path(root, filename=filename, work_item_id=work_item_id)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(payload, indent=2, sort_keys=True), encoding="utf-8")
    return ToolResult(
        call_id=call_id,
        status="ok",
        content=content,
        artifact_refs=[
            ArtifactRef(
                uri=str(path),
                type="dataset",
                metadata={
                    "filename": path.name,
                    "artifact_kind": artifact_kind,
                    "work_item_id": work_item_id,
                    **_artifact_count_metadata(payload),
                },
            )
        ],
        metadata={"path": str(path), "artifact_kind": artifact_kind, **_artifact_count_metadata(payload)},
    )


def _artifact_count_metadata(payload: dict[str, Any]) -> dict[str, Any]:
    keys = ("file_count", "source_count", "table_count", "row_count", "record_count", "accepted_count", "rejected_count")
    return {key: payload[key] for key in keys if isinstance(payload.get(key), int)}


def _artifact_path(root: Path, *, filename: str, work_item_id: str | None = None) -> Path:
    if work_item_id:
        return root / _safe_filename(work_item_id) / _safe_filename(filename)
    return root / _safe_filename(filename)


def _write_jsonl(path: Path, records: list[dict[str, Any]]) -> None:
    lines = [json.dumps(record, sort_keys=True) for record in records]
    path.write_text("\n".join(lines) + ("\n" if lines else ""), encoding="utf-8")


def _load_source_payload(arguments: dict[str, Any]) -> dict[str, Any]:
    if arguments.get("candidate_source_files_path"):
        return _load_json(_required_path(arguments, "candidate_source_files_path"))
    if arguments.get("document_inventory_path"):
        return _load_json(_required_path(arguments, "document_inventory_path"))
    raise ValueError("candidate_source_files_path or document_inventory_path is required")


def _candidate_source_items(payload: dict[str, Any]) -> list[dict[str, Any]]:
    value = payload.get("candidate_sources")
    if isinstance(value, list):
        return [item for item in value if isinstance(item, dict)]
    files = payload.get("files")
    if isinstance(files, list):
        return [item for item in files if isinstance(item, dict)]
    return []


def _source_ref(arguments: dict[str, Any]) -> str | None:
    for key in ("candidate_source_files_path", "document_inventory_path"):
        value = arguments.get(key)
        if isinstance(value, str):
            return value
    return None


def _source_path(source: dict[str, Any]) -> Path | None:
    value = source.get("path")
    if not isinstance(value, str) or not value:
        return None
    return Path(value).expanduser()


def _source_file_from_table(table: dict[str, Any]) -> Path | None:
    value = table.get("source_file")
    if not isinstance(value, str) or not value:
        return None
    return Path(value).expanduser()


def _table_reason(headers: Any) -> str:
    header_text = " ".join(_string_list(headers)).casefold()
    if any(token in header_text for token in ("sequence", "seq", "dna", "nucleotide")):
        return "table has sequence-like column labels"
    if any(token in header_text for token in ("name", "id", "component", "value", "activity", "score")):
        return "table has structured scientific record-like column labels"
    return "table is a structured source"


def _source_row_index(table: dict[str, Any], row_offset: int) -> int:
    header_row_index = table.get("header_row_index")
    if isinstance(header_row_index, int):
        return header_row_index + 1 + row_offset
    return row_offset


def _task_terms(task_goal: Any, schema_path: Any) -> set[str]:
    terms = {"sequence", "seq", "dna", "name", "id", "component", "record", "value", "evidence", "source"}
    if isinstance(task_goal, str):
        terms.update(_meaningful_tokens(task_goal))
    if isinstance(schema_path, str) and schema_path:
        try:
            terms.update(_meaningful_tokens(Path(schema_path).expanduser().read_text(encoding="utf-8")))
        except OSError:
            pass
    return {term for term in terms if len(term) >= 2}


def _meaningful_tokens(text: str) -> set[str]:
    stop = {
        "the",
        "and",
        "for",
        "from",
        "with",
        "that",
        "this",
        "when",
        "into",
        "are",
        "was",
        "were",
        "should",
    }
    return {token for token in re.findall(r"[A-Za-z][A-Za-z0-9_]{2,}", text.casefold()) if token not in stop}


def _iter_leaf_values(value: Any) -> list[Any]:
    if isinstance(value, dict):
        leaves = []
        for item in value.values():
            leaves.extend(_iter_leaf_values(item))
        return leaves
    if isinstance(value, list):
        leaves = []
        for item in value:
            leaves.extend(_iter_leaf_values(item))
        return leaves
    return [value]


def _normalize_sequence(value: Any) -> str:
    if value is None:
        return ""
    return "".join(str(value).upper().split())


def _is_dna_like(value: str, *, min_length: int) -> bool:
    normalized = _normalize_sequence(value)
    return len(normalized) >= min_length and set(normalized) <= {"A", "C", "G", "T", "N"}


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = "".join(char if char.isalnum() else "_" for char in text)
    text = "_".join(part for part in text.split("_") if part)
    return text or "column"


def _list_row(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, tuple):
        return list(value)
    if isinstance(value, dict):
        return list(value.values())
    return [value]


def _source_files(value: Any, *, root: Path) -> list[Path]:
    if not isinstance(value, list):
        return []
    paths = []
    for item in value:
        if not isinstance(item, str) or not item:
            continue
        path = Path(item).expanduser()
        if not path.is_absolute():
            path = root / path
        paths.append(path)
    return paths


def _required_path(arguments: dict[str, Any], key: str) -> Path:
    value = arguments.get(key)
    if not isinstance(value, str) or not value:
        raise ValueError(f"{key} must be a non-empty string")
    return Path(value).expanduser()


def _load_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _optional_string(value: Any) -> str | None:
    if isinstance(value, str) and value.strip():
        return value.strip()
    return None


def _optional_int(value: Any, *, default: int) -> int:
    if isinstance(value, bool):
        return default
    if isinstance(value, int):
        return value
    if isinstance(value, str) and value.isdigit():
        return int(value)
    return default


def _string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value if str(item)]
    return []


def _relative_path(path: Path, root: Path) -> str:
    try:
        return path.relative_to(root).as_posix()
    except ValueError:
        return path.as_posix()


def _safe_filename(value: str) -> str:
    safe = "".join(char if char.isalnum() or char in ("-", "_", ".") else "_" for char in value)
    return safe if safe and safe not in {".", ".."} else "artifact"


def _spec(name: str, description: str, properties: dict[str, Any], *, required: list[str] | None = None) -> ToolSpec:
    return ToolSpec(
        name=name,
        description=description,
        parameters_schema={"type": "object", "properties": properties, "required": required or []},
        metadata={"scientific_artifact_tool": True},
    )


def _build_document_inventory_spec() -> ToolSpec:
    return _spec(
        "build_document_inventory",
        "Scan a scientific work-item directory and write document_inventory.json.",
        {
            "root": {"type": "string"},
            "work_item_id": {"type": "string"},
            "source_files": {"type": "array", "items": {"type": "string"}},
        },
        required=["root"],
    )


def _discover_candidate_source_files_spec() -> ToolSpec:
    return _spec(
        "discover_candidate_source_files",
        "Rank inventory files likely to contain structured extractable scientific information.",
        {
            "document_inventory_path": {"type": "string"},
            "work_item_id": {"type": "string"},
            "task_goal": {"type": "string"},
            "schema_path": {"type": "string"},
            "max_sources": {"type": "integer", "minimum": 1},
        },
        required=["document_inventory_path"],
    )


def _discover_candidate_tables_spec() -> ToolSpec:
    return _spec(
        "discover_candidate_tables",
        "Inspect candidate source files and write candidate_tables.json with sheet/table metadata.",
        {
            "candidate_source_files_path": {"type": "string"},
            "document_inventory_path": {"type": "string"},
            "work_item_id": {"type": "string"},
            "max_tables": {"type": "integer", "minimum": 1},
            "max_sample_rows": {"type": "integer", "minimum": 1},
        },
    )


def _extract_candidate_rows_spec() -> ToolSpec:
    return _spec(
        "extract_candidate_rows",
        "Read candidate tables and write candidate_rows.json with sequence/value/name/evidence fields.",
        {
            "candidate_tables_path": {"type": "string"},
            "work_item_id": {"type": "string"},
            "schema_path": {"type": "string"},
            "max_rows_per_table": {"type": "integer", "minimum": 1},
            "min_sequence_length": {"type": "integer", "minimum": 1},
            "sequence_extraction_profile": {"type": "string"},
            "primary_component_tables_only": {"type": "boolean"},
        },
        required=["candidate_tables_path"],
    )


def _build_candidate_records_spec() -> ToolSpec:
    return _spec(
        "build_candidate_records",
        "Build schema-shaped candidate records from candidate_rows.json with source provenance.",
        {
            "candidate_rows_path": {"type": "string"},
            "schema_path": {"type": "string"},
            "article_id": {"type": "string"},
            "work_item_id": {"type": "string"},
            "sequence_extraction_profile": {"type": "string"},
            "deduplicate_sequences": {"type": "boolean"},
        },
        required=["candidate_rows_path"],
    )


def _validate_candidate_records_spec() -> ToolSpec:
    return _spec(
        "validate_candidate_records",
        "Validate candidate records into accepted/rejected records without using ground truth.",
        {
            "candidate_records_path": {"type": "string"},
            "schema_path": {"type": "string"},
            "work_item_id": {"type": "string"},
            "sequence_extraction_profile": {"type": "string"},
        },
        required=["candidate_records_path"],
    )


def _serialize_final_records_spec() -> ToolSpec:
    return _spec(
        "serialize_final_records",
        "Serialize candidate or validated records into final JSONL artifacts.",
        {
            "records_path": {"type": "string"},
            "records": {"type": "array", "items": {"type": "object"}},
            "artifact_name": {"type": "string"},
            "also_write_final_records": {"type": "boolean"},
            "work_item_id": {"type": "string"},
        },
    )


def _register_if_missing(registry: ToolRegistry, spec: ToolSpec, handler: Any) -> None:
    if registry.get_spec(spec.name) is None:
        registry.register(spec, handler)
