from __future__ import annotations

import csv
import json
from pathlib import Path
import re
import statistics
from typing import Any
from zipfile import ZipFile
import xml.etree.ElementTree as ET

from evolab.contracts.tools import ToolResult, ToolSpec
from evolab.tools.paths import resolve_path_arguments
from evolab.tools.runtime import ToolRegistry


def register_table_tools(registry: ToolRegistry, *, base_dir: str | Path | None = None) -> None:
    for spec, handler in [
        (_inspect_table_spec(), inspect_table),
        (_read_table_slice_spec(), read_table_slice),
        (_inspect_excel_workbook_spec(), inspect_excel_workbook),
        (_read_excel_sheet_spec(), read_excel_sheet),
        (_detect_table_header_spec(), detect_table_header),
        (_normalize_table_spec(), normalize_table),
        (_profile_table_spec(), profile_table),
    ]:
        _register_if_missing(
            registry,
            spec,
            lambda arguments, handler=handler: handler(
                resolve_path_arguments(arguments, base_dir=base_dir, names=("path",))
            ),
        )


def inspect_table(arguments: dict[str, Any]) -> ToolResult:
    table = _load_table(arguments)
    rows = table["rows"]
    header_result = _detect_header(rows)
    headers = header_result["headers"]
    preview_count = int(arguments.get("max_rows_preview", 20))
    payload = {
        "path": table.get("path"),
        "sheet_name": table.get("sheet_name"),
        "row_count": len(rows),
        "column_count": _column_count(rows),
        "headers": headers,
        "preview_rows": rows[:preview_count],
        "warnings": [*table.get("warnings", []), *header_result.get("warnings", [])],
    }
    if table.get("plain_text_table_block") is not None:
        payload["plain_text_table_block"] = table["plain_text_table_block"]
    return _result("inspect_table", payload, f"inspected table with {payload['row_count']} rows")


def read_table_slice(arguments: dict[str, Any]) -> ToolResult:
    table = _load_table(arguments)
    rows = table["rows"]
    start_row = int(arguments.get("start_row", 0))
    end_row = arguments.get("end_row")
    end = int(end_row) if end_row is not None else len(rows)
    warnings = list(table.get("warnings", []))
    start_row, end, source_line_warning = _normalize_plain_text_slice_bounds(
        table=table,
        row_count=len(rows),
        start_row=start_row,
        end_row=end,
    )
    if source_line_warning:
        warnings.append(source_line_warning)
    selected = rows[start_row:end]
    columns = arguments.get("columns")
    if isinstance(columns, list) and columns:
        selected = [_select_columns(row, columns) for row in selected]
    header_result = _detect_header(rows)
    payload = {
        "path": table.get("path"),
        "sheet_name": table.get("sheet_name"),
        "start_row": start_row,
        "end_row": end,
        "headers": header_result["headers"],
        "rows": selected,
        "warnings": [*warnings, *header_result.get("warnings", [])],
    }
    if table.get("plain_text_table_block") is not None:
        payload["plain_text_table_block"] = table["plain_text_table_block"]
    return _result("read_table_slice", payload, f"read {len(selected)} table rows")


def _normalize_plain_text_slice_bounds(
    *,
    table: dict[str, Any],
    row_count: int,
    start_row: int,
    end_row: int,
) -> tuple[int, int, str | None]:
    block = table.get("plain_text_table_block")
    if not isinstance(block, dict):
        return start_row, end_row, None
    source_start = block.get("start_line")
    source_end = block.get("end_line")
    if not isinstance(source_start, int) or not isinstance(source_end, int):
        return start_row, end_row, None
    if start_row < row_count and end_row <= row_count:
        return start_row, end_row, None
    if end_row <= source_start or start_row > source_end:
        return start_row, end_row, None

    relative_start = max(0, start_row - source_start)
    relative_end = min(row_count, max(relative_start, end_row - source_start))
    return (
        relative_start,
        relative_end,
        "interpreted start_row/end_row as source line indexes",
    )


def inspect_excel_workbook(arguments: dict[str, Any]) -> ToolResult:
    path = _path_argument(arguments)
    workbook = _load_openpyxl_workbook(path)
    sheets = []
    for sheet in workbook.worksheets:
        rows = _sheet_rows(sheet)
        header_result = _detect_header(rows)
        sheets.append(
            {
                "sheet_name": sheet.title,
                "max_row": sheet.max_row,
                "max_column": sheet.max_column,
                "non_empty_rows": len([row for row in rows if any(_non_empty(cell) for cell in row)]),
                "headers": header_result["headers"],
            }
        )
    return _result(
        "inspect_excel_workbook",
        {"path": str(path), "sheet_count": len(sheets), "sheets": sheets},
        f"inspected workbook with {len(sheets)} sheets",
    )


def read_excel_sheet(arguments: dict[str, Any]) -> ToolResult:
    path = _path_argument(arguments)
    sheet_name = arguments.get("sheet_name")
    workbook = _load_openpyxl_workbook(path)
    sheet = workbook[str(sheet_name)] if isinstance(sheet_name, str) and sheet_name else workbook.worksheets[0]
    rows = _sheet_rows(sheet)
    start_row = int(arguments.get("start_row", 0))
    end_row = arguments.get("end_row")
    end = int(end_row) if end_row is not None else len(rows)
    selected = rows[start_row:end]
    columns = arguments.get("columns")
    if isinstance(columns, list) and columns:
        selected = [_select_columns(row, columns) for row in selected]
    header_result = _detect_header(rows)
    payload = {
        "path": str(path),
        "sheet_name": sheet.title,
        "dimensions": {"rows": sheet.max_row, "columns": sheet.max_column},
        "headers": header_result["headers"],
        "rows": selected,
    }
    return _result("read_excel_sheet", payload, f"read {len(selected)} rows from {sheet.title}")


def detect_table_header(arguments: dict[str, Any]) -> ToolResult:
    rows = _rows_from_arguments(arguments)
    payload = _detect_header(rows)
    return _result("detect_table_header", payload, f"detected header row {payload['header_row_index']}")


def normalize_table(arguments: dict[str, Any]) -> ToolResult:
    rows = _rows_from_arguments(arguments)
    headers = _string_list(arguments.get("headers"))
    if not headers:
        header_result = _detect_header(rows)
        headers = header_result["headers"]
        if header_result["header_row_index"] is not None:
            rows = rows[int(header_result["header_row_index"]) + 1 :]
    rows = [_normalize_row(row) for row in rows]
    rows = [row for row in rows if any(_non_empty(cell) for cell in row)]
    column_count = max([len(headers), *[len(row) for row in rows]], default=0)
    empty_columns = {
        index
        for index in range(column_count)
        if all(index >= len(row) or not _non_empty(row[index]) for row in rows)
        and (index >= len(headers) or not _non_empty(headers[index]))
    }
    normalized_headers = [_normalize_header(headers[index] if index < len(headers) else f"column_{index + 1}") for index in range(column_count) if index not in empty_columns]
    normalized_rows = [
        [row[index] if index < len(row) else "" for index in range(column_count) if index not in empty_columns]
        for row in rows
    ]
    payload = {"headers": normalized_headers, "rows": normalized_rows, "row_count": len(normalized_rows), "column_count": len(normalized_headers)}
    return _result("normalize_table", payload, f"normalized table to {len(normalized_rows)} rows")


def profile_table(arguments: dict[str, Any]) -> ToolResult:
    normalized = normalize_table(arguments).metadata
    headers = normalized["headers"]
    rows = normalized["rows"]
    profiles = []
    for index, header in enumerate(headers):
        values = [str(row[index]) for row in rows if index < len(row) and _non_empty(row[index])]
        lengths = [len(value) for value in values]
        profiles.append(
            {
                "name": header,
                "non_empty_count": len(values),
                "unique_count": len(set(values)),
                "sample_values": values[:5],
                "numeric_fraction": _fraction(values, _is_number),
                "dna_like_fraction": _fraction(values, _is_dna_like),
                "avg_length": round(statistics.mean(lengths), 3) if lengths else 0.0,
            }
        )
    payload = {"headers": headers, "row_count": len(rows), "column_profiles": profiles, "warnings": []}
    return _result("profile_table", payload, f"profiled {len(profiles)} columns")


def _load_table(arguments: dict[str, Any]) -> dict[str, Any]:
    path = _path_argument(arguments)
    suffix = path.suffix.casefold()
    if suffix == ".csv":
        rows = _delimited_rows(path, arguments.get("delimiter") or ",")
    elif suffix == ".tsv":
        rows = _delimited_rows(path, arguments.get("delimiter") or "\t")
    elif suffix in {".md", ".markdown"}:
        markdown = _markdown_table(path, arguments)
        rows = markdown["rows"]
        return {
            "path": str(path),
            "sheet_name": arguments.get("sheet_name"),
            "rows": rows,
            "warnings": markdown.get("warnings", []),
            "plain_text_table_block": markdown.get("plain_text_table_block"),
        }
    elif suffix == ".xlsx":
        workbook = _load_openpyxl_workbook(path)
        sheet_name = arguments.get("sheet_name")
        sheet = workbook[str(sheet_name)] if isinstance(sheet_name, str) and sheet_name else workbook.worksheets[0]
        rows = _sheet_rows(sheet)
        return {"path": str(path), "sheet_name": sheet.title, "rows": rows, "warnings": []}
    else:
        rows = _delimited_rows(path, arguments.get("delimiter") or ",")
    return {"path": str(path), "sheet_name": arguments.get("sheet_name"), "rows": rows, "warnings": []}


def _rows_from_arguments(arguments: dict[str, Any]) -> list[list[Any]]:
    rows = arguments.get("rows")
    if isinstance(rows, list):
        return [_list_row(row) for row in rows]
    return _load_table(arguments)["rows"]


def _delimited_rows(path: Path, delimiter: Any) -> list[list[str]]:
    with path.open(newline="", encoding="utf-8") as handle:
        return [list(row) for row in csv.reader(handle, delimiter=str(delimiter))]


def _markdown_table(path: Path, arguments: dict[str, Any]) -> dict[str, Any]:
    table_caption = arguments.get("table_caption")
    if isinstance(table_caption, str) and table_caption.strip():
        return _plain_text_markdown_table(path, table_caption.strip())
    pipe_rows = _pipe_markdown_rows(path)
    if pipe_rows:
        return {"rows": pipe_rows, "warnings": []}
    return _plain_text_markdown_table(path, None)


def _pipe_markdown_rows(path: Path) -> list[list[str]]:
    rows = []
    for line in path.read_text(encoding="utf-8").splitlines():
        stripped = line.strip()
        if not stripped.startswith("|") or not stripped.endswith("|"):
            continue
        cells = [cell.strip() for cell in stripped.strip("|").split("|")]
        if cells and all(set(cell) <= {"-", ":", " "} for cell in cells):
            continue
        rows.append(cells)
    return rows


def _plain_text_markdown_table(path: Path, table_caption: str | None) -> dict[str, Any]:
    lines = path.read_text(encoding="utf-8").splitlines()
    start_index = _plain_table_start_index(lines, table_caption)
    if start_index is None:
        detail = f"caption {table_caption!r} not found" if table_caption else "no plain text table caption found"
        return {"rows": [], "warnings": [detail]}
    end_index = _plain_table_end_index(lines, start_index)
    block_lines = [
        line.strip()
        for line in lines[start_index + 1 : end_index]
        if _plain_table_content_line(line)
    ]
    rows = _plain_table_rows(block_lines)
    caption = lines[start_index].strip()
    return {
        "rows": rows,
        "warnings": [] if rows else ["plain text table block did not contain parseable rows"],
        "plain_text_table_block": {
            "caption": caption,
            "start_line": start_index + 1,
            "end_line": end_index,
            "source_format": "plain_text_markdown",
        },
    }


def _plain_table_start_index(lines: list[str], table_caption: str | None) -> int | None:
    if table_caption:
        needle = table_caption.casefold()
        for index, line in enumerate(lines):
            if needle in line.casefold():
                return index
        return None
    for index, line in enumerate(lines):
        if re.match(r"^\s*Table\s+\S+", line, flags=re.IGNORECASE):
            return index
    return None


def _plain_table_end_index(lines: list[str], start_index: int) -> int:
    for index in range(start_index + 1, len(lines)):
        line = lines[index].strip()
        if re.match(r"^Table\s+\S+", line, flags=re.IGNORECASE):
            return index
        if re.match(r"^Figure\s+\S+", line, flags=re.IGNORECASE):
            return index
    return len(lines)


def _plain_table_content_line(line: str) -> bool:
    stripped = line.strip()
    if not stripped:
        return False
    if stripped.startswith("<!--") and stripped.endswith("-->"):
        return False
    if re.fullmatch(r"S\d{2,}", stripped):
        return False
    return True


def _plain_table_rows(lines: list[str]) -> list[list[str]]:
    first_label_index = next(
        (index for index, line in enumerate(lines) if _looks_like_plain_row_label(line)),
        None,
    )
    if first_label_index is None:
        return []
    headers = [_normalize_header(line) for line in lines[:first_label_index] if line.strip()]
    data_rows = _plain_data_rows(lines[first_label_index:])
    if not headers:
        return data_rows
    return [headers, *data_rows]


def _plain_data_rows(lines: list[str]) -> list[list[str]]:
    rows: list[list[str]] = []
    label: str | None = None
    dna_parts: list[str] = []
    values: list[str] = []
    extras: list[str] = []

    def flush() -> None:
        nonlocal label, dna_parts, values, extras
        if label is None:
            return
        row = [label]
        if dna_parts:
            row.append("".join(dna_parts))
        row.extend(values)
        row.extend(extras)
        rows.append(row)
        label = None
        dna_parts = []
        values = []
        extras = []

    for raw_line in lines:
        line = raw_line.strip()
        if _looks_like_plain_row_label(line):
            flush()
            label = line
            continue
        if label is None:
            continue
        if _is_dna_like(line):
            dna_parts.append(line.upper().replace(" ", ""))
        elif _is_number(line):
            values.append(line)
        else:
            extras.append(line)
    flush()
    return rows


def _looks_like_plain_row_label(value: str) -> bool:
    stripped = value.strip()
    if not stripped or len(stripped) > 32:
        return False
    if _is_number(stripped) or _is_dna_like(stripped):
        return False
    return re.fullmatch(r"[A-Za-z]{0,8}\d+[A-Za-z0-9_.-]*", stripped) is not None


def _load_openpyxl_workbook(path: Path) -> Any:
    try:
        from openpyxl import load_workbook
    except ModuleNotFoundError:
        return _load_simple_xlsx_workbook(path)
    return load_workbook(path, read_only=True, data_only=True)


def _sheet_rows(sheet: Any) -> list[list[Any]]:
    return [[cell for cell in row] for row in sheet.iter_rows(values_only=True)]


class _SimpleWorkbook:
    def __init__(self, worksheets: list[Any]) -> None:
        self.worksheets = worksheets

    def __getitem__(self, name: str) -> Any:
        for sheet in self.worksheets:
            if sheet.title == name:
                return sheet
        raise KeyError(name)


class _SimpleSheet:
    def __init__(self, title: str, rows: list[list[Any]]) -> None:
        self.title = title
        self._rows = rows
        self.max_row = len(rows)
        self.max_column = _column_count(rows)

    def iter_rows(self, values_only: bool = True) -> list[list[Any]]:
        return self._rows


def _load_simple_xlsx_workbook(path: Path) -> _SimpleWorkbook:
    with ZipFile(path) as archive:
        shared_strings = _xlsx_shared_strings(archive)
        sheet_targets = _xlsx_sheet_targets(archive)
        sheets = [
            _SimpleSheet(name, _xlsx_sheet_rows(archive, target, shared_strings))
            for name, target in sheet_targets
        ]
    return _SimpleWorkbook(sheets)


def _xlsx_shared_strings(archive: ZipFile) -> list[str]:
    if "xl/sharedStrings.xml" not in archive.namelist():
        return []
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    strings = []
    for item in root.findall(".//{*}si"):
        strings.append("".join(text.text or "" for text in item.findall(".//{*}t")))
    return strings


def _xlsx_sheet_targets(archive: ZipFile) -> list[tuple[str, str]]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    rels = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    rel_by_id = {
        rel.attrib["Id"]: rel.attrib["Target"]
        for rel in rels.findall("{*}Relationship")
        if "Id" in rel.attrib and "Target" in rel.attrib
    }
    sheets = []
    for sheet in workbook.findall(".//{*}sheet"):
        name = sheet.attrib.get("name", "Sheet")
        rel_id = sheet.attrib.get("{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id")
        target = rel_by_id.get(str(rel_id))
        if target:
            sheets.append((name, _xlsx_target_path(target)))
    return sheets


def _xlsx_target_path(target: str) -> str:
    normalized = target.lstrip("/")
    if normalized.startswith("xl/"):
        return normalized
    return f"xl/{normalized}"


def _xlsx_sheet_rows(archive: ZipFile, target: str, shared_strings: list[str]) -> list[list[Any]]:
    root = ET.fromstring(archive.read(target))
    rows = []
    for row in root.findall(".//{*}sheetData/{*}row"):
        values: list[Any] = []
        for cell in row.findall("{*}c"):
            index = _xlsx_cell_column_index(cell.attrib.get("r", ""))
            while len(values) <= index:
                values.append("")
            values[index] = _xlsx_cell_value(cell, shared_strings)
        rows.append(values)
    return rows


def _xlsx_cell_column_index(reference: str) -> int:
    letters = "".join(char for char in reference if char.isalpha()).upper()
    index = 0
    for char in letters:
        index = index * 26 + (ord(char) - ord("A") + 1)
    return max(0, index - 1)


def _xlsx_cell_value(cell: ET.Element, shared_strings: list[str]) -> Any:
    value = cell.find("{*}v")
    if value is None:
        inline = cell.find(".//{*}t")
        return inline.text if inline is not None else ""
    text = value.text or ""
    if cell.attrib.get("t") == "s":
        try:
            return shared_strings[int(text)]
        except (ValueError, IndexError):
            return text
    return text


def _detect_header(rows: list[list[Any]]) -> dict[str, Any]:
    best_index: int | None = None
    best_score = -1.0
    warnings = []
    for index, row in enumerate(rows[:20]):
        non_empty = [cell for cell in row if _non_empty(cell)]
        string_count = sum(isinstance(cell, str) for cell in non_empty)
        score = string_count + min(len(non_empty), 4) * 0.25
        if len(non_empty) <= 1:
            score -= 2.0
        if score > best_score:
            best_index = index
            best_score = score
    if best_index is None:
        return {"header_row_index": None, "headers": [], "confidence": 0.0, "warnings": ["no rows available"]}
    headers = [_normalize_header(str(cell)) for cell in rows[best_index]]
    if not headers:
        warnings.append("selected header row is empty")
    confidence = min(1.0, max(0.1, best_score / 6.0))
    return {"header_row_index": best_index, "headers": headers, "confidence": round(confidence, 3), "warnings": warnings}


def _select_columns(row: list[Any], columns: list[Any]) -> list[Any]:
    indexes = [int(column) for column in columns if isinstance(column, int) or (isinstance(column, str) and column.isdigit())]
    return [row[index] if index < len(row) else "" for index in indexes]


def _list_row(value: Any) -> list[Any]:
    if isinstance(value, list):
        return value
    if isinstance(value, dict):
        return list(value.values())
    return [value]


def _normalize_row(row: list[Any]) -> list[str]:
    return ["" if cell is None else str(cell).strip() for cell in row]


def _normalize_header(value: Any) -> str:
    text = str(value or "").strip().casefold()
    text = "".join(char if char.isalnum() else "_" for char in text)
    text = "_".join(part for part in text.split("_") if part)
    return text or "column"


def _column_count(rows: list[list[Any]]) -> int:
    return max((len(row) for row in rows), default=0)


def _non_empty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def _is_number(value: str) -> bool:
    try:
        float(value)
        return True
    except ValueError:
        return False


def _is_dna_like(value: str) -> bool:
    stripped = value.strip().upper()
    return len(stripped) >= 8 and set(stripped) <= {"A", "C", "G", "T", "N"}


def _fraction(values: list[str], predicate: Any) -> float:
    if not values:
        return 0.0
    return round(sum(1 for value in values if predicate(value)) / len(values), 3)


def _path_argument(arguments: dict[str, Any]) -> Path:
    path = arguments.get("path")
    if not isinstance(path, str) or not path:
        raise ValueError("path must be a non-empty string")
    return Path(path).expanduser()


def _string_list(value: Any) -> list[str]:
    if isinstance(value, list):
        return [str(item) for item in value]
    return []


def _result(call_id: str, payload: dict[str, Any], content: str) -> ToolResult:
    return ToolResult(call_id=call_id, status="ok", content=content, metadata=payload)


def _spec(
    name: str,
    description: str,
    properties: dict[str, Any],
    *,
    required: list[str] | None = None,
) -> ToolSpec:
    return ToolSpec(
        name=name,
        description=description,
        parameters_schema={
            "type": "object",
            "properties": properties,
            "required": required or [],
        },
    )


def _path_properties() -> dict[str, Any]:
    return {
        "path": {"type": "string", "description": "Path to a CSV, TSV, markdown, or XLSX table."},
        "delimiter": {"type": "string", "description": "Delimiter for delimited text tables."},
        "sheet_name": {"type": "string", "description": "Optional sheet name for XLSX workbooks."},
        "table_caption": {
            "type": "string",
            "description": "Optional markdown/plain-text table caption to locate, such as 'Table S1'.",
        },
    }


def _rows_property() -> dict[str, Any]:
    return {
        "rows": {
            "type": "array",
            "items": {
                "type": "array",
                "items": {"type": "string", "description": "Cell value serialized as text."},
            },
            "description": "Table rows as arrays of cell values.",
        }
    }


def _slice_properties() -> dict[str, Any]:
    return {
        **_path_properties(),
        "start_row": {"type": "integer", "minimum": 0, "description": "Zero-based first row to include."},
        "end_row": {"type": "integer", "minimum": 0, "description": "Exclusive row index where reading stops."},
        "columns": {
            "type": "array",
            "items": {"type": "integer", "description": "Column index."},
            "description": "Optional list of zero-based column indexes to include.",
        },
    }


def _inspect_table_spec() -> ToolSpec:
    return _spec(
        "inspect_table",
        "Inspect CSV, TSV, markdown, or XLSX table structure.",
        {
            **_path_properties(),
            "max_rows_preview": {
                "type": "integer",
                "minimum": 1,
                "description": "Maximum preview rows to return.",
            },
        },
        required=["path"],
    )


def _read_table_slice_spec() -> ToolSpec:
    return _spec(
        "read_table_slice",
        "Read a row and column slice from a table artifact.",
        _slice_properties(),
        required=["path"],
    )


def _inspect_excel_workbook_spec() -> ToolSpec:
    return _spec(
        "inspect_excel_workbook",
        "Inspect workbook sheets and dimensions.",
        {"path": {"type": "string", "description": "Path to an XLSX workbook."}},
        required=["path"],
    )


def _read_excel_sheet_spec() -> ToolSpec:
    return _spec(
        "read_excel_sheet",
        "Read rows from one Excel workbook sheet.",
        {
            "path": {"type": "string", "description": "Path to an XLSX workbook."},
            "sheet_name": {"type": "string", "description": "Optional sheet name. Defaults to the first sheet."},
            "start_row": {"type": "integer", "minimum": 0, "description": "Zero-based first row to include."},
            "end_row": {"type": "integer", "minimum": 0, "description": "Exclusive row index where reading stops."},
            "columns": {
                "type": "array",
                "items": {"type": "integer", "description": "Column index."},
                "description": "Optional list of zero-based column indexes to include.",
            },
        },
        required=["path"],
    )


def _detect_table_header_spec() -> ToolSpec:
    return _spec(
        "detect_table_header",
        "Detect likely table header row.",
        _rows_property(),
    )


def _normalize_table_spec() -> ToolSpec:
    return _spec(
        "normalize_table",
        "Normalize headers and remove empty table rows or columns.",
        {
            **_path_properties(),
            **_rows_property(),
            "headers": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Optional header names to use.",
            },
        },
    )


def _profile_table_spec() -> ToolSpec:
    return _spec(
        "profile_table",
        "Profile table columns with generic string and numeric statistics.",
        {
            **_path_properties(),
            **_rows_property(),
            "headers": {
                "type": "array",
                "items": {"type": "string"},
                "description": "Optional header names to use.",
            },
        },
    )


def _register_if_missing(registry: ToolRegistry, spec: ToolSpec, handler: Any) -> None:
    if registry.get_spec(spec.name) is None:
        registry.register(spec, handler)
